"""
pymake
-------------------------------
 - Eugenio Marinetto
 - nenetto@gmail.com
-------------------------------
Created 29-05-2018
"""
import pandas as pd
from pymake.main import printer as pm
from pymake.utils.pandas.functions import normalize_str_nosp
import numpy as np
from openpyxl import load_workbook
import os
import tempfile
import shutil


def iter_rows(ws):
    for row in ws.iter_rows():
        yield [cell.value for cell in row[0:3]]


def get_schema(fname):
    wb = load_workbook(fname) #from openpyxl import load_workbook
    ws = wb.active
    schema = pd.DataFrame(iter_rows(ws))
    schema.columns = schema.iloc[0]
    schema = schema.reindex(schema.index.drop(0))
    return schema


def separate_numeric_column(df, column_name, verbose=True):
    """
    This function try to convert a column to a numeric. Those values where a number is not found are set to nans

    :param df: dataframe for input
    :type df: pandas.DataFrame
    :param column_name: name of the column to be fixed
    :type column_name: str
    """
    pm.print_info('Fixing column {0}'.format(column_name))
    if verbose:
        pm.print_info('Number of rows {0}'.format(df.shape[0]))

    # Copy column of interest
    dfx = df[[column_name]].copy()

    # Create type variable
    dfx[column_name + '_type'] = 'num'

    n_num = 0
    n_str = 0
    n_nans = 0
    n_others = 0

    total = dfx.shape[0]
    for i, row in dfx.iterrows():
        pm.print_info_percentage(100*i/total, 'Processing column', padding=1)

        try: # Try conversion to number
            x = float(row[column_name])

            if np.isnan(x):
                n_nans += 1

            n_num += 1

        except ValueError:
            dfx.loc[i, column_name + '_type'] = row[column_name]
            dfx.loc[i, column_name] = np.nan
            n_str += 1
        except TypeError:
            n_others += 1
            pass # This check for timeseries and date types

    pm.print_info_percentage(100, 'Processed  column', padding=1)

    pm.print_info('Nums: {0}'.format(n_num))
    pm.print_info('Nans: {0}'.format(n_nans))
    pm.print_info('Strs: {0}'.format(n_str))
    pm.print_info('Unkn: {0}'.format(n_others))

    if verbose:
        pm.print_info('Number of different types reduces from {0} to {1}'.format(len(df[column_name].unique()),
                                                                                 len(dfx[column_name + '_type'].unique())))

        pm.print_info('Classification:')
        for e in dfx[column_name + '_type'].unique():

            n = dfx[dfx[column_name + '_type'] == e].shape[0]

            pm.print_info_2('{0} # {1}'.format(e, n), padding=1)

    # check the number of non numeric values

    total_num = dfx[column_name].shape[0]

    if total_num == (n_num + n_nans):
        # Numeric variable - leave as it
        pm.print_warning('Seems to be numeric, please revise')
        pass
    elif total_num == n_str:
        # Categorical variable - leave as it
        pm.print_warning('Seems to be categorical, please revise')
        pass
    elif total_num == n_others:
        # Unknown type or date - leave as it
        pm.print_warning('Unknown or date, please revise')
        pass
    elif n_str > (n_num + n_nans):
        # Categorical variable - leave as it
        pm.print_warning('Seems to be categorical, please revise')
    else:
        # Mixed variable, do the split in two
        df[column_name] = dfx[column_name].copy().astype('float')
        df[column_name + '_type'] = dfx[column_name + '_type'].copy().astype('str')


def fixtable(df):

    # Rename column names
    new_cs = list()
    for c in df.columns:
        new_c = normalize_str_nosp(c)

        if new_c in new_cs:
            new_c += '_x'

        new_cs.append(new_c)

    df.columns = new_cs

    # Separate numeric columns if there exists
    for c in df.columns:
        separate_numeric_column(df, c, verbose=False)

    # Reorganize columns
    new_c = list()
    used_c = list()

    for c in df.columns:

        if (c + '_type') in df.columns:
            if (c + '_type') not in used_c:
                new_c.append(c)
                new_c.append(c + '_type')
                used_c.append(c + '_type')
        else:
            if c not in used_c:
                new_c.append(c)

    newdf = df[new_c].copy()

    return newdf


def summary_table(df, fixedtable_file_xlsx, schema_file_xlsx, summary_file_xlsx):

    try:
        from pydqc.infer_schema import infer_schema
        from pydqc.data_summary import data_summary

    except ImportError:
        pm.print_error('To use this function, you need to install pacakge pydqc')
        pm.print_error('    - https://github.com/nenetto/pydqc')
        pm.print_error('    - pip[3] install https://github.com/nenetto/pydqc')
        pm.print_error('', exit_code=1)

    # Fix table
    df = fixtable(df)
    df.to_excel(fixedtable_file_xlsx, index=False)

    dirpath = tempfile.mkdtemp()

    # Infer schema
    pm.print_info('Infering Schema')
    infer_schema(df, fname='',
                 output_root=dirpath,
                 sample_size=1.0,
                 type_threshold=0.5,
                 n_jobs=1,
                 base_schema=None)

    shutil.copyfile(os.path.join(dirpath, 'data_schema_.xlsx'), schema_file_xlsx)
    df_schema = get_schema(schema_file_xlsx)

    pm.print_info('Generating Summary')
    data_summary(table_schema=df_schema, table=df, output_root=dirpath, fname='', sample_size=1.0, keep_images=False)
    shutil.copyfile(os.path.join(dirpath, 'data_summary_.xlsx'), summary_file_xlsx)

    # Remove temp dir
    shutil.rmtree(dirpath)
